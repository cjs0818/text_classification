# -*- coding: utf-8 -*-

#------------------------
#  Required in python2.7
#import sys
#reload(sys)
#sys.setdefaultencoding('utf-8')
#------------------------

import json
import os.path
from openpyxl import load_workbook as xl

File_Name = 'callerReceiverPair.xlsx'        # xlsx name
xlsx_Path = "./"   # xlsx_Path

def write_data(i, data):
    wb = xl(xlsx_Path + File_Name)
    ws = wb.active
    ws['D'+str(i)] = data
    wb.save(xlsx_Path+File_Name)

def read_xfile(num): #Insert path & line number
    wb = xl(xlsx_Path+File_Name)

    #shList = wb.get_sheet_names()   # Deprecated
    shList = wb.sheetnames

    #sh = wb.get_sheet_by_name(shList[0])  # Deprecated
    sh = wb[shList[0]]

    try:
        c1 = int(sh['C'+str(num)].value[0:8])#sh['C'+str(num)].value[0:15]
        c2 = int(sh['C'+str(num)].value[10:15])
        t1  = int(sh['B'+str(num)].value[0:8])#sh['B'+str(num)].value[0:15]
        t2 = int(sh['B' + str(num)].value[10:15])
        #print c1, c2, t1,t2
        if c1-t1 is 0 and c2-t2 is 0:
            caller = sh['C' + str(num)].value[0:28]
            taker = sh['B' + str(num)].value[0:28]
        else :
            err_msg = 'Error, contents mismatch'
            print(err_msg)
            write_data(num, err_msg)
            caller = 'contents'
            taker = 'contents'
    except:
        err_msg = 'Error, no name'
        print(err_msg)
        write_data(num, err_msg)
        caller = 'no'
        taker = 'no'

    return caller, taker


def read_text(filepath, num, name):
    data = None
    if os.path.isfile(filepath+".TEXT"):
        with open(filepath+".TEXT", encoding="utf8") as f:

            try :
                data=json.load(f)
            except:
                err_msg = 'Error, data character'
                print(err_msg)
                write_data(num, err_msg)
                return 0
    elif name is 'contents':
        return 0
    elif name is 'no':
        return 0
    else:
        err_msg = 'Error, File name mismatch'
        print(err_msg, filepath)
        write_data(num, err_msg)
        return 0
    return data

def read_situation(filepath, num, name):
    data = None
    if os.path.isfile(filepath+".SITUATION"):
        with open(filepath+".SITUATION", encoding="utf8") as f:

            try :
                data=json.load(f)
            except:
                err_msg = 'Error, data character'
                print(err_msg)
                write_data(num, err_msg)
                return 0
    elif name is 'contents':
        return 0
    elif name is 'no':
        return 0
    else:
        err_msg = 'Error, File name mismatch'
        print(err_msg, filepath)
        write_data(num, err_msg)
        return 0
    return data

def read_situation_corection(filepath, num, name):
    data = None
    if os.path.isfile(filepath+".SITUATION"):
        with open(filepath+".SITUATION", encoding="utf8") as f:

            lines = ''
            cnt = 0
            for line_before in f:
                line_after = line_before.rstrip()
                if cnt == 2:
                    line_after = line_after + ','
                lines = lines + line_after + '\n'
                cnt += 1
            #print(lines)
            return lines


    elif name is 'contents':
        return 0
    elif name is 'no':
        return 0
    else:
        err_msg = 'Error, File name mismatch'
        print(err_msg, filepath)
        write_data(num, err_msg)
        return 0
    return data


def read_truelevel_corection(filepath, num, name):
    data = None
    if os.path.isfile(filepath+".TRUE_LEVEL"):
        with open(filepath+".TRUE_LEVEL", encoding="utf8") as f:

            lines = ''
            cnt = 0
            for line_before in f:
                line_after = line_before.rstrip()
                if cnt == 2:
                    line_after = line_after + ','
                lines = lines + line_after + '\n'
                cnt += 1
            #print(lines)
            return lines


    elif name is 'contents':
        return 0
    elif name is 'no':
        return 0
    else:
        err_msg = 'Error, File name mismatch'
        print(err_msg, filepath)
        write_data(num, err_msg)
        return 0
    return data




def write_text(num, msg, idx, start):
    if idx == start:
        write_type = "w"
    else:
        write_type = "a"

    for i in range(num):
        if msg[i][0] is 'caller':
            with open(xlsx_Path+"sentence_caller.txt", write_type, encoding="utf8") as f:
                f.write(str(msg[i][1])+"\n")
        else:
            with open(xlsx_Path+"sentence_taker.txt", write_type, encoding="utf8") as f:
                f.write(str(msg[i][1])+"\n")

def write_sentences(sentences, idx, start):
    if idx == start:
        write_type = "w"
    else:
        write_type = "a"

    with open(xlsx_Path + "listen_situation.txt", write_type, encoding="utf8") as f:
        f.write(str(sentences) + "\n")

def write_situation_corection(filepath, sentences):
    with open(filepath + ".SITUATION", "w", encoding="utf8") as f:
        f.write(str(sentences))

def write_truelevel_corection(filepath, sentences):
    with open(filepath + ".TRUE_LEVEL", "w", encoding="utf8") as f:
        f.write(str(sentences))
